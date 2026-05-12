"""
Writes a new 'Scenarios_Updated' sheet into a COPY of the original xlsx. The
original sheets ('AEIS', 'Comps') are passed through untouched.

The sensitivity grid is plain formulas (`=$A2*B$1`), not an openpyxl Data
Table object. Data Tables can corrupt the file when openpyxl reads/writes —
we leave native interactivity to the user (Data > What-If Analysis) if they
want it.

Polish layer applied at the end of `write`:
  * Row 2 callout (merged, light gray)
  * 3-color conditional format scale on the sensitivity grid
  * Explicit scenario reference rows below the grid
  * "Upside vs current" header tied to D10 (no hardcoded prices)
  * `<ignoredErrors sqref="D8"/>` injected into the saved xml so Excel does
    not show the inconsistent-formula warning triangle on the PW EV cell.
    openpyxl does not serialize this element, so we post-process the xml.
"""
from __future__ import annotations

import shutil
import zipfile
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .scenarios import (
    ScenarioBundle, eps_grid, pe_grid, what_changed_summary,
    CURRENT_TRADING_EPS, CURRENT_TRADING_PE,
)


HEADER_FILL = PatternFill("solid", fgColor="305496")  # dark blue
HEADER_FONT = Font(bold=True, color="FFFFFF")
SECTION_FONT = Font(bold=True, size=12)
# Light gray callout fill used to draw the eye to the PW EV row.
EV_HIGHLIGHT_FILL = PatternFill("solid", fgColor="F2F2F2")
THIN = Side(border_style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _set_header(ws: Worksheet, cell: str, text: str) -> None:
    """Style a single cell as a column header (dark fill, white bold text)."""
    ws[cell] = text
    ws[cell].fill = HEADER_FILL
    ws[cell].font = HEADER_FONT
    ws[cell].alignment = Alignment(horizontal="center", vertical="center")


def _scenario_block(ws: Worksheet, top_row: int, bundle: ScenarioBundle) -> None:
    """Write the bear/base/bull block. Upside formulas reference $D$10."""
    ws.cell(row=top_row, column=1, value="Updated bear / base / bull").font = SECTION_FONT
    headers = ["Scenario", "EPS (CY27)", "P/E", "Price",
               "Upside vs current", "Probability"]
    for i, h in enumerate(headers, start=1):
        _set_header(ws, f"{get_column_letter(i)}{top_row+1}", h)

    # Scenario rows render with a plain white background; the scenario name
    # in column A is the only bolded cell. Color cues live in the sensitivity
    # grid gradient and the EV highlight below.
    for offset, sc in enumerate(
        (bundle.bear, bundle.base, bundle.bull), start=top_row + 2
    ):
        ws.cell(row=offset, column=1, value=sc.name)
        ws.cell(row=offset, column=2, value=sc.eps).number_format = '"$"#,##0.00'
        ws.cell(row=offset, column=3, value=sc.pe).number_format = '0.0"x"'
        ws.cell(row=offset, column=4, value=f"=B{offset}*C{offset}")
        ws.cell(row=offset, column=4).number_format = '"$"#,##0.00'
        ws.cell(row=offset, column=5, value=f"=D{offset}/$D$10-1")
        ws.cell(row=offset, column=5).number_format = "0.0%"
        ws.cell(row=offset, column=6, value=sc.probability).number_format = "0%"
        for col in range(1, 7):
            cell = ws.cell(row=offset, column=col)
            cell.border = BOX
            if col == 1:
                cell.font = Font(bold=True)

    # Expected value row (explicit weighted sum keeps Excel's
    # inconsistent-formula warning from flagging a SUMPRODUCT against the
    # adjacent =B*C cells; the IgnoredErrors XML injection below is the
    # belt-and-suspenders fix.) Light gray fill highlights this as the key
    # output line.
    ev_row = top_row + 5
    ws.cell(row=ev_row, column=1, value="Prob-weighted EV").font = Font(bold=True)
    ws.cell(
        row=ev_row, column=4,
        value=(f"=D{top_row+2}*F{top_row+2}"
               f"+D{top_row+3}*F{top_row+3}"
               f"+D{top_row+4}*F{top_row+4}"),
    )
    ws.cell(row=ev_row, column=4).number_format = '"$"#,##0.00'
    ws.cell(row=ev_row, column=4).font = Font(bold=True)
    ws.cell(row=ev_row, column=5, value=f"=D{ev_row}/$D$10-1")
    ws.cell(row=ev_row, column=5).number_format = "0.0%"
    ws.cell(row=ev_row, column=5).font = Font(bold=True)
    for col in range(1, 7):
        ws.cell(row=ev_row, column=col).fill = EV_HIGHLIGHT_FILL

    # Current implied anchor
    anchor_row = ev_row + 1
    ws.cell(row=anchor_row, column=1,
            value="Current implied (consensus EPS x current ~32x)").font = Font(italic=True)
    ws.cell(row=anchor_row, column=2,
            value=CURRENT_TRADING_EPS).number_format = '"$"#,##0.00'
    ws.cell(row=anchor_row, column=3,
            value=CURRENT_TRADING_PE).number_format = '0.0"x"'
    ws.cell(row=anchor_row, column=4,
            value=f"=B{anchor_row}*C{anchor_row}").number_format = '"$"#,##0.00'
    ws.cell(row=anchor_row, column=5,
            value=f"=D{anchor_row}/$D$10-1").number_format = "0.0%"

    # Current price (the anchor that all upside formulas reference)
    cp_row = anchor_row + 1
    ws.cell(row=cp_row, column=1, value="Current price (May 4)").font = Font(italic=True)
    ws.cell(row=cp_row, column=4,
            value=bundle.current_price).number_format = '"$"#,##0.00'


def _callout_row(ws: Worksheet, row: int, text: str) -> None:
    """Bold light-gray summary callout merged across A:F."""
    for rng in list(ws.merged_cells.ranges):
        if str(rng).startswith(f"A{row}:"):
            ws.unmerge_cells(str(rng))
    ws.merge_cells(f"A{row}:F{row}")
    cell = ws[f"A{row}"]
    cell.value = text
    cell.font = Font(bold=True, size=11)
    cell.fill = PatternFill("solid", fgColor="E7E6E6")
    cell.alignment = Alignment(
        horizontal="left", vertical="center", wrap_text=True
    )
    ws.row_dimensions[row].height = 38


def _conditional_color_scale(ws: Worksheet, sqref: str) -> None:
    """
    Monochrome blue 3-color scale (white -> light blue -> dark navy).
    Buy side appropriate palette; the heat-map structure is unchanged.
    Applied to every conditionally formatted range for consistency.
    """
    rule = ColorScaleRule(
        start_type="min", start_color="FFFFFF",
        mid_type="percentile", mid_value=50, mid_color="9DC3E6",
        end_type="max", end_color="1F4E79",
    )
    ws.conditional_formatting.add(sqref, rule)


def _scenario_reference_rows(ws: Worksheet, top_row: int,
                              bundle: ScenarioBundle) -> None:
    """
    Compact bear/base/bull reference table below the sensitivity grid so the
    PM can see exact scenario points without eyeballing the grid.
    """
    ws.cell(row=top_row, column=1,
            value="Scenario point reference (exact)").font = Font(bold=True, size=12)

    headers = ["", "EPS", "P/E", "Price", "Upside vs current"]
    for i, h in enumerate(headers, start=1):
        if not h:
            continue
        c = ws.cell(row=top_row, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")

    rows = [
        ("Bear scenario", bundle.bear.eps, bundle.bear.pe),
        ("Base scenario", bundle.base.eps, bundle.base.pe),
        ("Bull scenario", bundle.bull.eps, bundle.bull.pe),
    ]
    for offset, (name, eps, pe) in enumerate(rows, start=top_row + 1):
        ws.cell(row=offset, column=1, value=name).font = Font(bold=True)
        ws.cell(row=offset, column=2, value=eps).number_format = '"$"#,##0.00'
        ws.cell(row=offset, column=3, value=pe).number_format = '0.0"x"'
        ws.cell(row=offset, column=4,
                value=f"=B{offset}*C{offset}").number_format = '"$"#,##0.00'
        ws.cell(row=offset, column=5,
                value=f"=D{offset}/$D$10-1").number_format = "0.0%"
        for col in "ABCDE":
            ws[f"{col}{offset}"].border = BOX


def _inject_ignored_errors(xlsx_path: Path, sheet_name: str,
                           sqref: str) -> None:
    """
    Post-process the saved xlsx to add <ignoredErrors sqref="D8" formula="1"/>.
    openpyxl does not serialize this element, so we rewrite the sheet xml.
    The element must come after pageMargins/pageSetup per ECMA-376.
    """
    # Resolve sheet xml path by looking up workbook.xml.rels
    tmp = xlsx_path.with_suffix(".tmp.xlsx")
    injection = (
        f"<ignoredErrors>"
        f'<ignoredError sqref="{sqref}" formula="1"/>'
        f"</ignoredErrors>"
    )

    # Identify which sheetN.xml hosts our named sheet by reading workbook.xml.
    # Attribute order is not guaranteed, so search the whole element then pull
    # attrs out of the match.
    import re as _re
    with zipfile.ZipFile(xlsx_path, "r") as z:
        workbook_xml = z.read("xl/workbook.xml").decode("utf-8")
        rels_xml = z.read("xl/_rels/workbook.xml.rels").decode("utf-8")
    sheet_el = _re.search(
        rf'<sheet[^>]*name="{_re.escape(sheet_name)}"[^>]*/>',
        workbook_xml,
    )
    if not sheet_el:
        return
    rid_match = _re.search(r'r:id="([^"]+)"', sheet_el.group(0))
    if not rid_match:
        return
    rid = rid_match.group(1)
    rel_el = _re.search(
        rf'<Relationship[^>]*Id="{_re.escape(rid)}"[^>]*/>',
        rels_xml,
    )
    if not rel_el:
        return
    target_match = _re.search(r'Target="([^"]+)"', rel_el.group(0))
    if not target_match:
        return
    sheet_path = "xl/" + target_match.group(1).lstrip("/").removeprefix("xl/")

    with zipfile.ZipFile(xlsx_path, "r") as zin:
        with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == sheet_path:
                    xml = data.decode("utf-8")
                    if "<ignoredErrors>" not in xml:
                        xml = xml.replace(
                            "</worksheet>", injection + "</worksheet>", 1
                        )
                    data = xml.encode("utf-8")
                zout.writestr(item, data)
    shutil.move(str(tmp), str(xlsx_path))


def _changed_block(ws: Worksheet, top_row: int, summary: dict) -> None:
    """Old vs new diff."""
    ws.cell(row=top_row, column=1,
            value="What changed vs May 1 model").font = SECTION_FONT
    headers = ["Scenario", "Old EPS", "Old P/E", "Old price",
               "New EPS", "New P/E", "New price"]
    for i, h in enumerate(headers, start=1):
        _set_header(ws, f"{get_column_letter(i)}{top_row+1}", h)
    rows = [
        ("Bear", summary["eps_old"], summary["pe_old_bear"],
         summary["old_scenarios"]["bear"],
         summary["eps_new_bear"], summary["pe_new_bear"],
         summary["new_scenarios"]["bear"]),
        ("Base", summary["eps_old"], summary["pe_old_base"],
         summary["old_scenarios"]["base"],
         summary["eps_new_base"], summary["pe_new_base"],
         summary["new_scenarios"]["base"]),
        ("Bull", summary["eps_old"], summary["pe_old_bull"],
         summary["old_scenarios"]["bull"],
         summary["eps_new_bull"], summary["pe_new_bull"],
         summary["new_scenarios"]["bull"]),
    ]
    for off, row in enumerate(rows, start=top_row + 2):
        ws.cell(row=off, column=1, value=row[0]).font = Font(bold=True)
        for col_idx, val in enumerate(row[1:], start=2):
            cell = ws.cell(row=off, column=col_idx, value=val)
            if col_idx in (3, 6):
                cell.number_format = '0.0"x"'
            elif col_idx in (2, 5):
                cell.number_format = '"$"#,##0.00'
            else:
                cell.number_format = '"$"#,##0.00'
            cell.border = BOX


def _sensitivity_grid(ws: Worksheet, top_row: int) -> None:
    """
    Plain-formula sensitivity grid.

    Layout (with top_row as the row of the header):
      - top_row, col A: blank
      - top_row, cols B..: P/E values across
      - top_row+1.., col A: EPS values down
      - Intersections: =$A{row}*B${top_row}

    No openpyxl Data Table object — formulas update on Excel recalc.
    """
    ws.cell(row=top_row, column=1, value="Sensitivity grid (EPS x P/E)").font = SECTION_FONT
    grid_top = top_row + 2  # header row for P/E
    grid_left = 1  # column A holds EPS

    eps_vals = eps_grid()
    pe_vals = pe_grid()

    # P/E header row (across)
    ws.cell(row=grid_top, column=grid_left, value="EPS \\ P/E").font = Font(bold=True)
    ws.cell(row=grid_top, column=grid_left).fill = HEADER_FILL
    ws.cell(row=grid_top, column=grid_left).font = HEADER_FONT
    for j, pe in enumerate(pe_vals, start=grid_left + 1):
        cell = ws.cell(row=grid_top, column=j, value=pe)
        cell.number_format = '0.0"x"'
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    # EPS column (down) + intersection formulas
    pe_header_letter_offsets = [get_column_letter(grid_left + 1 + j)
                                for j in range(len(pe_vals))]
    for i, eps in enumerate(eps_vals, start=grid_top + 1):
        eps_cell = ws.cell(row=i, column=grid_left, value=eps)
        eps_cell.number_format = '"$"#,##0.00'
        eps_cell.fill = HEADER_FILL
        eps_cell.font = HEADER_FONT
        eps_cell.alignment = Alignment(horizontal="center")
        for j_off, pe_letter in enumerate(pe_header_letter_offsets, start=grid_left + 1):
            eps_letter = get_column_letter(grid_left)
            formula = f"=${eps_letter}{i}*{pe_letter}${grid_top}"
            cell = ws.cell(row=i, column=j_off, value=formula)
            cell.number_format = '"$"#,##0'
            cell.border = BOX
            cell.alignment = Alignment(horizontal="center")


SHEET_NAME = "Scenarios_Updated"


def write(template_path: Path, output_path: Path,
          bundle: ScenarioBundle) -> None:
    """Copy the original xlsx and append a polished Scenarios_Updated sheet."""
    wb = load_workbook(template_path)
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
    ws = wb.create_sheet(SHEET_NAME)

    ws.column_dimensions["A"].width = 38
    for col in "BCDEFGHIJKLMNOPQRS":
        ws.column_dimensions[col].width = 14

    ws["A1"] = "AEIS Updated Scenarios (post Q1 2026 + BAC/Cowen revisions)"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:F1")

    ev_dir = "below" if bundle.expected_value_upside_pct < 0 else "above"
    _callout_row(
        ws, row=2,
        text=(
            f"PW EV ${bundle.expected_value:.0f} sits "
            f"{abs(bundle.expected_value_upside_pct):.0f}% {ev_dir} current "
            f"${bundle.current_price:.0f}. "
            f"Base case is roughly current price. "
            f"Stock pricing partial bull case. "
            f"Risk reward skewed negative: bear downside "
            f"{abs(bundle.bear.upside_pct(bundle.current_price)):.0f}%, "
            f"bull upside "
            f"{abs(bundle.bull.upside_pct(bundle.current_price)):.0f}%."
        ),
    )

    _scenario_block(ws, top_row=3, bundle=bundle)
    _changed_block(ws, top_row=14, summary=what_changed_summary())
    _sensitivity_grid(ws, top_row=21)
    _conditional_color_scale(ws, "B24:K32")
    _scenario_reference_rows(ws, top_row=33, bundle=bundle)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    # IgnoredErrors must be injected after openpyxl save since openpyxl does
    # not serialize that xml element on Worksheet.
    _inject_ignored_errors(output_path, SHEET_NAME, sqref="D8")
